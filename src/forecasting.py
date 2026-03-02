"""
forecasting.py — SKU bazlı talep tahmini.

REF-04 modelleri:
  Baseline  : Naive (son 7 gün ort.), Simple MA (30 gün), Weighted MA (ewm)
  Orta      : Holt-Winters (ETS additive) — duzgun/sezonluk SKU
  Aralıklı  : Croston / TSB — aralikli / toplu SKU
  Duzensiz  : Robust MA (kayan medyan) — duzensiz SKU
  Yeni      : Basit ortalama + geniş güven bandı

Akış:
  1. Her SKU için talep_tipi'ne göre model seç
  2. Backtesting: son BACKTEST_GUN gün test; MAPE/WAPE/MAE/bias hesapla
  3. Tam veri üzerinde yeniden fit → TAHMIN_HORIZONLAR günlük tahmin üret
  4. T3 formatında DataFrame döndür

Çıktı (T3): barkod, tarih, tahmin_adet, alt_band, ust_band, model, tahmin_horizon
"""

import logging
import sys
import warnings
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.config import (
    DATA_PROCESSED_DIR,
    LOG_FORMAT,
    LOG_LEVEL,
    TAHMIN_HORIZONLAR,
)

logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT)
log = logging.getLogger(__name__)

# ─── Sabitler ────────────────────────────────────────────────────────────────
BACKTEST_GUN: int = 30          # Test seti uzunluğu (gün)
BAND_Z_VARSAYILAN: float = 1.65  # %90 güven aralığı (yeni/az veri → 2.0 kullanılır)
BAND_Z_GENIS: float = 2.0       # Geniş band (yeni SKU)
MA_PENCERE: int = 30            # Baseline MA penceresi
EWM_SPAN: int = 14              # Exponentially weighted MA span
MIN_TRAIN_GUN: int = 14         # Model fit için minimum eğitim günü
FORECAST_PARQUET = DATA_PROCESSED_DIR / "forecast_results.parquet"


# ─── Metrik Yardımcıları ─────────────────────────────────────────────────────

def _mape(gercek: np.ndarray, tahmin: np.ndarray) -> float:
    """Mean Absolute Percentage Error (sıfır gerçek değerler atlanır)."""
    mask = gercek > 0
    if mask.sum() == 0:
        return np.nan
    return float(np.mean(np.abs((gercek[mask] - tahmin[mask]) / gercek[mask])) * 100)


def _wape(gercek: np.ndarray, tahmin: np.ndarray) -> float:
    """Weighted Absolute Percentage Error (model seçiminde kullanılır)."""
    toplam = gercek.sum()
    if toplam == 0:
        return np.nan
    return float(np.abs(gercek - tahmin).sum() / toplam * 100)


def _mae(gercek: np.ndarray, tahmin: np.ndarray) -> float:
    return float(np.mean(np.abs(gercek - tahmin)))


def _bias(gercek: np.ndarray, tahmin: np.ndarray) -> float:
    """Pozitif → fazla tahmin, negatif → eksik tahmin."""
    return float(np.mean(tahmin - gercek))


def _metrikler(gercek: np.ndarray, tahmin: np.ndarray) -> dict:
    return {
        "mape": _mape(gercek, tahmin),
        "wape": _wape(gercek, tahmin),
        "mae": _mae(gercek, tahmin),
        "bias": _bias(gercek, tahmin),
    }


# ─── Model Fonksiyonları ──────────────────────────────────────────────────────

def _naive_tahmin(seri: np.ndarray, horizon: int) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Son 7 günün ortalamasını horizon boyunca yayar."""
    pencere = min(7, len(seri))
    ort = np.mean(seri[-pencere:])
    std = np.std(seri[-pencere:]) if pencere > 1 else ort * 0.5
    tahmin = np.full(horizon, max(ort, 0.0))
    alt = np.maximum(tahmin - BAND_Z_VARSAYILAN * std, 0)
    ust = tahmin + BAND_Z_VARSAYILAN * std
    return tahmin, alt, ust


def _simple_ma_tahmin(seri: np.ndarray, horizon: int, pencere: int = MA_PENCERE) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Basit kayan ortalama (son `pencere` gün)."""
    pencere = min(pencere, len(seri))
    ort = np.mean(seri[-pencere:])
    std = np.std(seri[-pencere:]) if pencere > 1 else ort * 0.5
    tahmin = np.full(horizon, max(ort, 0.0))
    alt = np.maximum(tahmin - BAND_Z_VARSAYILAN * std, 0)
    ust = tahmin + BAND_Z_VARSAYILAN * std
    return tahmin, alt, ust


def _ewm_tahmin(seri: np.ndarray, horizon: int) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Exponentially weighted mean (span=EWM_SPAN)."""
    s = pd.Series(seri)
    ewm_val = float(s.ewm(span=min(EWM_SPAN, len(seri)), adjust=False).mean().iloc[-1])
    ewm_val = max(ewm_val, 0.0)
    std = float(s.std()) if len(s) > 1 else ewm_val * 0.5
    tahmin = np.full(horizon, ewm_val)
    alt = np.maximum(tahmin - BAND_Z_VARSAYILAN * std, 0)
    ust = tahmin + BAND_Z_VARSAYILAN * std
    return tahmin, alt, ust


def _robust_ma_tahmin(seri: np.ndarray, horizon: int) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Kayan medyan (duzensiz talep için daha dayanıklı)."""
    pencere = min(MA_PENCERE, len(seri))
    medyan = float(np.median(seri[-pencere:]))
    mad = float(np.median(np.abs(seri[-pencere:] - medyan)))  # Median Absolute Deviation
    medyan = max(medyan, 0.0)
    tahmin = np.full(horizon, medyan)
    alt = np.maximum(tahmin - BAND_Z_VARSAYILAN * mad, 0)
    ust = tahmin + BAND_Z_VARSAYILAN * mad * 2  # Asimetrik band (duzensiz skewness)
    return tahmin, alt, ust


def _yeni_sku_tahmin(seri: np.ndarray, horizon: int) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Az veri SKU: basit ortalama + geniş güven bandı."""
    # VARSAYIM: Yeni/az veri SKU'larında geniş belirsizlik bandı (z=2.0)
    ort = max(np.mean(seri), 0.0) if len(seri) > 0 else 0.0
    std = float(np.std(seri)) if len(seri) > 1 else ort * 0.8
    tahmin = np.full(horizon, ort)
    alt = np.maximum(tahmin - BAND_Z_GENIS * std, 0)
    ust = tahmin + BAND_Z_GENIS * std
    return tahmin, alt, ust


def _croston_tahmin(seri: np.ndarray, horizon: int, alfa: float = 0.1) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    Croston'ın yöntemi: talep miktarı ve talep aralığını ayrı ayrı düzleştirir.

    Args:
        seri: Günlük adet (0 dahil).
        horizon: Tahmin günü.
        alfa: Düzleştirme parametresi (0 < alfa < 1).

    Returns:
        (tahmin, alt_band, ust_band) — her biri `horizon` uzunluklu array.
    """
    pozitif_idx = np.where(seri > 0)[0]
    if len(pozitif_idx) < 2:
        return _yeni_sku_tahmin(seri, horizon)

    # Başlangıç değerleri
    d_hat = float(seri[pozitif_idx[0]])  # talep büyüklüğü tahmini
    p_hat = float(pozitif_idx[0] + 1)   # talep aralığı tahmini (gün)

    for i in range(1, len(pozitif_idx)):
        adet = float(seri[pozitif_idx[i]])
        aralik = float(pozitif_idx[i] - pozitif_idx[i - 1])
        d_hat = alfa * adet + (1 - alfa) * d_hat
        p_hat = alfa * aralik + (1 - alfa) * p_hat

    croston_ort = max(d_hat / max(p_hat, 1.0), 0.0)

    # Güven bandı: talep büyüklüklerinin std'si / aralık
    talep_buyuklukleri = seri[pozitif_idx].astype(float)
    std = float(np.std(talep_buyuklukleri)) / max(p_hat, 1.0)

    tahmin = np.full(horizon, croston_ort)
    alt = np.maximum(tahmin - BAND_Z_VARSAYILAN * std, 0)
    ust = tahmin + BAND_Z_VARSAYILAN * std
    return tahmin, alt, ust


def _tsb_tahmin(seri: np.ndarray, horizon: int, alfa: float = 0.1, beta: float = 0.1) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    TSB (Teunter-Syntetos-Babai) yöntemi.
    Croston'a göre daha iyi: talep olasılığını da düzleştirir.

    Args:
        seri: Günlük adet (0 dahil).
        horizon: Tahmin günü.
        alfa: Talep büyüklüğü düzleştirme parametresi.
        beta: Talep olasılığı düzleştirme parametresi.

    Returns:
        (tahmin, alt_band, ust_band) — her biri `horizon` uzunluklu array.
    """
    pozitif_idx = np.where(seri > 0)[0]
    if len(pozitif_idx) < 2:
        return _yeni_sku_tahmin(seri, horizon)

    n = len(seri)
    d_hat = float(seri[pozitif_idx[0]])
    p_hat = len(pozitif_idx) / n  # başlangıç olasılığı

    talep_buyuklukleri = []
    for i in range(n):
        if seri[i] > 0:
            d_hat = alfa * seri[i] + (1 - alfa) * d_hat
            p_hat = beta * 1 + (1 - beta) * p_hat
            talep_buyuklukleri.append(seri[i])
        else:
            p_hat = beta * 0 + (1 - beta) * p_hat

    tsb_ort = max(d_hat * p_hat, 0.0)
    std = float(np.std(talep_buyuklukleri)) * p_hat if talep_buyuklukleri else tsb_ort * 0.5

    tahmin = np.full(horizon, tsb_ort)
    alt = np.maximum(tahmin - BAND_Z_VARSAYILAN * std, 0)
    ust = tahmin + BAND_Z_VARSAYILAN * std
    return tahmin, alt, ust


def _holt_winters_tahmin(seri: np.ndarray, horizon: int) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    Holt-Winters (additive) — duzgun talep için.
    Mevsimsel periyot: 7 (haftalık).
    statsmodels.ExponentialSmoothing kullanır.
    """
    from statsmodels.tsa.holtwinters import ExponentialSmoothing

    n = len(seri)
    # Mevsimsel periyot 7 için en az 2 tam periyot (14 gün) gerekli
    if n < 14:
        return _simple_ma_tahmin(seri, horizon)

    # Sıfır değerleri küçük bir sayıya çevir (log dönüşüm hatalarını önler)
    seri_s = np.where(seri == 0, 0.01, seri.astype(float))

    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            # Sezonluk periyot: veri yeterliyse 7 gün, değilse trend+level only
            sezonluk = "add" if n >= 28 else None
            mevsim_periyot = 7 if sezonluk else None

            model = ExponentialSmoothing(
                seri_s,
                trend="add",
                seasonal=sezonluk,
                seasonal_periods=mevsim_periyot,
                initialization_method="estimated",
            )
            fit = model.fit(optimized=True, use_brute=False)

        tahmin_raw = fit.forecast(horizon)
        tahmin_raw = np.maximum(tahmin_raw, 0.0)

        # Güven bandı: rezidüel std kullanarak
        reziduel_std = float(np.std(fit.resid))
        alt = np.maximum(tahmin_raw - BAND_Z_VARSAYILAN * reziduel_std, 0)
        ust = tahmin_raw + BAND_Z_VARSAYILAN * reziduel_std
        return tahmin_raw, alt, ust

    except Exception as exc:
        log.debug("HoltWinters hata → Simple MA fallback: %s", exc)
        return _simple_ma_tahmin(seri, horizon)


# ─── Model Seçim & Yürütme ───────────────────────────────────────────────────

_MODEL_FON = {
    "naive": _naive_tahmin,
    "simple_ma": _simple_ma_tahmin,
    "ewm": _ewm_tahmin,
    "robust_ma": _robust_ma_tahmin,
    "holt_winters": _holt_winters_tahmin,
    "croston": _croston_tahmin,
    "tsb": _tsb_tahmin,
    "yeni_sku": _yeni_sku_tahmin,
}


def _talep_tipine_gore_modeller(talep_tipi: str) -> list[str]:
    """Talep tipine göre sıralı aday model listesi döndürür (en iyi → fallback)."""
    _MAP = {
        "duzgun": ["holt_winters", "ewm", "simple_ma"],
        "aralikli": ["croston", "tsb", "simple_ma"],
        "toplu": ["tsb", "croston", "robust_ma"],
        "duzensiz": ["robust_ma", "ewm", "simple_ma"],
        "yeni": ["yeni_sku", "simple_ma"],
    }
    return _MAP.get(talep_tipi, ["simple_ma"])


def _en_iyi_modeli_sec(
    seri_train: np.ndarray,
    seri_test: np.ndarray,
    aday_modeller: list[str],
) -> tuple[str, dict]:
    """
    Verilen aday modeller içinden WAPE bazlı en iyi modeli seçer.

    Args:
        seri_train: Eğitim serisi.
        seri_test: Test serisi (son BACKTEST_GUN gün).
        aday_modeller: Denenecek model isimleri.

    Returns:
        (en_iyi_model_adi, metrik_dict)
    """
    horizon = len(seri_test)
    en_iyi_model = aday_modeller[0]
    en_iyi_wape = np.inf
    en_iyi_metrik: dict = {}

    for model_adi in aday_modeller:
        try:
            fon = _MODEL_FON[model_adi]
            tahmin_arr, _, _ = fon(seri_train, horizon)
            metrik = _metrikler(seri_test, tahmin_arr)
            wape_val = metrik["wape"]
            if not np.isnan(wape_val) and wape_val < en_iyi_wape:
                en_iyi_wape = wape_val
                en_iyi_model = model_adi
                en_iyi_metrik = metrik
        except Exception as exc:
            log.debug("Model '%s' hata: %s", model_adi, exc)

    if not en_iyi_metrik:
        # Hiç model çalışmadıysa basit MA fallback
        tahmin_arr, _, _ = _simple_ma_tahmin(seri_train, horizon)
        en_iyi_metrik = _metrikler(seri_test, tahmin_arr)
        en_iyi_model = "simple_ma"

    return en_iyi_model, en_iyi_metrik


# ─── SKU Tahmin Motoru ────────────────────────────────────────────────────────

def _sku_tahmin(
    barkod: str,
    daily: pd.DataFrame,
    talep_tipi: str,
) -> list[dict]:
    """
    Tek bir SKU için backtesting + tam tahmin üretir.

    Returns:
        T3 formatında satır listesi.
    """
    sku_daily = daily[daily["barkod"] == barkod].sort_values("tarih")
    seri = sku_daily["adet"].to_numpy(dtype=float)
    tarihler = sku_daily["tarih"].values

    if len(seri) < MIN_TRAIN_GUN:
        log.debug("SKU %s: yetersiz veri (%d gün), atlanıyor.", barkod, len(seri))
        return []

    # Eğitim / test bölünmesi
    test_uzunluk = min(BACKTEST_GUN, len(seri) // 3)  # En fazla verinin 1/3'ü
    train_seri = seri[:-test_uzunluk]
    test_seri = seri[-test_uzunluk:]

    if len(train_seri) < MIN_TRAIN_GUN:
        train_seri = seri
        test_seri = seri[-7:]

    # Model seç (WAPE bazlı)
    aday_modeller = _talep_tipine_gore_modeller(talep_tipi)
    en_iyi_model, backtest_metrik = _en_iyi_modeli_sec(train_seri, test_seri, aday_modeller)

    log.debug(
        "SKU %s | tip=%s | model=%s | WAPE=%.1f%%",
        barkod, talep_tipi, en_iyi_model,
        backtest_metrik.get("wape", float("nan")),
    )

    # Tam veri üzerinde tahmin üret
    fon = _MODEL_FON[en_iyi_model]
    son_tarih = pd.Timestamp(tarihler[-1])
    satirlar: list[dict] = []

    for horizon in TAHMIN_HORIZONLAR:
        try:
            tahmin_arr, alt_arr, ust_arr = fon(seri, horizon)
        except Exception as exc:
            log.debug("Tahmin hatası %s/%d: %s", barkod, horizon, exc)
            tahmin_arr, alt_arr, ust_arr = _simple_ma_tahmin(seri, horizon)

        for gun_idx in range(horizon):
            tahmin_tarihi = son_tarih + pd.Timedelta(days=gun_idx + 1)
            satirlar.append({
                "barkod": barkod,
                "tarih": tahmin_tarihi,
                "tahmin_adet": round(float(tahmin_arr[gun_idx]), 4),
                "alt_band": round(float(alt_arr[gun_idx]), 4),
                "ust_band": round(float(ust_arr[gun_idx]), 4),
                "model": en_iyi_model,
                "tahmin_horizon": horizon,
                # Backtest metrikleri (bilgi amaçlı)
                "backtest_wape": round(backtest_metrik.get("wape", float("nan")), 2),
                "backtest_mape": round(backtest_metrik.get("mape", float("nan")), 2),
                "backtest_mae": round(backtest_metrik.get("mae", float("nan")), 4),
                "backtest_bias": round(backtest_metrik.get("bias", float("nan")), 4),
            })

    return satirlar


# ─── Ana Fonksiyon ───────────────────────────────────────────────────────────

def tahmin_yap(
    master: pd.DataFrame,
    daily: pd.DataFrame,
    kaydet: bool = True,
) -> pd.DataFrame:
    """
    Tüm SKU'lar için tahmin üretir. T3 formatında DataFrame döndürür.

    Args:
        master: Ürün master (talep_tipi sütunu gerekli).
        daily:  Günlük talep DataFrame (tarih normalize edilmiş).
        kaydet: True ise parquet kaydeder.

    Returns:
        T3 — barkod, tarih, tahmin_adet, alt_band, ust_band, model, tahmin_horizon
    """
    log.info("Tahmin başlıyor. SKU: %d | Horizon: %s gün", len(master), TAHMIN_HORIZONLAR)

    tum_satirlar: list[dict] = []
    for _, satir in master.iterrows():
        barkod = satir["barkod"]
        talep_tipi = satir.get("talep_tipi", "yeni")
        try:
            satirlar = _sku_tahmin(barkod, daily, talep_tipi)
            tum_satirlar.extend(satirlar)
        except Exception as exc:
            log.warning("SKU %s tahmin hatası (atlandı): %s", barkod, exc)

    if not tum_satirlar:
        log.error("Hiçbir SKU için tahmin üretilemedi.")
        return pd.DataFrame()

    sonuc = pd.DataFrame(tum_satirlar)
    sonuc["tarih"] = pd.to_datetime(sonuc["tarih"])
    sonuc = sonuc.sort_values(["barkod", "tahmin_horizon", "tarih"])

    log.info(
        "Tahmin tamamlandı. Toplam satır: %d | SKU: %d | Model dağılımı:\n%s",
        len(sonuc),
        sonuc["barkod"].nunique(),
        sonuc.drop_duplicates("barkod")["model"].value_counts().to_string(),
    )

    if kaydet:
        DATA_PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
        sonuc.to_parquet(FORECAST_PARQUET, index=False)
        log.info("Kaydedildi: %s", FORECAST_PARQUET)

    return sonuc


def faz2_excel_yaz(forecast: pd.DataFrame, master: pd.DataFrame) -> Path:
    """
    Faz 2 tahmin sonuçlarını formatlı Excel'e yazar.

    Sayfalar:
      T3_Tahmin        — tüm tahminler (barkod × horizon × tarih)
      Backtest_Ozet    — SKU başına backtest metrikleri ve seçilen model
      Model_Dagilimi   — model bazında SKU sayısı ve ort. WAPE

    Args:
        forecast: tahmin_yap() çıktısı (T3 DataFrame).
        master:   Ürün master (kanonik_ad için).

    Returns:
        Yazılan dosyanın Path nesnesi.
    """
    from src.config import DATA_OUTPUTS_DIR
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    DATA_OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    cikti = DATA_OUTPUTS_DIR / "faz2_tahmin.xlsx"

    # Kanonik adı ekle
    ad_map = master.set_index("barkod")["kanonik_ad"].to_dict()
    forecast = forecast.copy()
    forecast.insert(1, "kanonik_ad", forecast["barkod"].map(ad_map))

    # T3 — tüm tahminler
    t3 = forecast.rename(columns={
        "barkod": "Barkod",
        "kanonik_ad": "Ürün Adı",
        "tarih": "Tahmin Tarihi",
        "tahmin_adet": "Tahmin Adet",
        "alt_band": "Alt Band",
        "ust_band": "Üst Band",
        "model": "Model",
        "tahmin_horizon": "Horizon (gün)",
        "backtest_wape": "WAPE (%)",
        "backtest_mape": "MAPE (%)",
        "backtest_mae": "MAE",
        "backtest_bias": "Bias",
    })

    # Backtest özeti — SKU başına (her horizon için ayrı satır değil, tek satır)
    backtest_ozet = (
        forecast.drop_duplicates(subset=["barkod", "tahmin_horizon"])
        [["barkod", "kanonik_ad", "model", "tahmin_horizon",
          "backtest_wape", "backtest_mape", "backtest_mae", "backtest_bias"]]
        .merge(master[["barkod", "abc_sinifi", "talep_tipi"]], on="barkod", how="left")
        .sort_values(["abc_sinifi", "backtest_wape"])
        .rename(columns={
            "barkod": "Barkod",
            "kanonik_ad": "Ürün Adı",
            "model": "Seçilen Model",
            "tahmin_horizon": "Horizon (gün)",
            "backtest_wape": "WAPE (%)",
            "backtest_mape": "MAPE (%)",
            "backtest_mae": "MAE",
            "backtest_bias": "Bias",
            "abc_sinifi": "ABC",
            "talep_tipi": "Talep Tipi",
        })
    )

    # Model dağılımı
    model_dag = (
        forecast.drop_duplicates("barkod")[["model", "barkod"]]
        .groupby("model")
        .agg(sku_sayisi=("barkod", "count"))
        .reset_index()
        .merge(
            forecast.groupby("model")["backtest_wape"].mean().reset_index().rename(
                columns={"backtest_wape": "Ort. WAPE (%)"}
            ),
            on="model",
        )
        .rename(columns={"model": "Model", "sku_sayisi": "SKU Sayısı"})
        .sort_values("Ort. WAPE (%)")
    )

    _HEADER_FILL = "1565C0"

    def _format_ws(ws):
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col_cells
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 40)
        ws.freeze_panes = "A2"

    log.info("Excel yazılıyor: %s", cikti)
    with pd.ExcelWriter(cikti, engine="openpyxl") as writer:
        t3.to_excel(writer, sheet_name="T3_Tahmin", index=False)
        backtest_ozet.to_excel(writer, sheet_name="Backtest_Ozet", index=False)
        model_dag.to_excel(writer, sheet_name="Model_Dagilimi", index=False)

        wb = writer.book
        for sayfa in ["T3_Tahmin", "Backtest_Ozet", "Model_Dagilimi"]:
            _format_ws(wb[sayfa])

    log.info("Faz 2 Excel kaydedildi: %s", cikti)
    return cikti


# ─── Script olarak çalıştırma ─────────────────────────────────────────────────
if __name__ == "__main__":
    from src.config import DAILY_DEMAND_PARQUET, PRODUCT_MASTER_PARQUET

    master = pd.read_parquet(PRODUCT_MASTER_PARQUET)
    daily = pd.read_parquet(DAILY_DEMAND_PARQUET)

    forecast = tahmin_yap(master, daily)

    print("\n=== Tahmin Özeti ===")
    print(f"Toplam satır  : {len(forecast):,}")
    print(f"Benzersiz SKU : {forecast['barkod'].nunique()}")
    print("\nModel dağılımı:")
    print(forecast.drop_duplicates("barkod")["model"].value_counts().to_string())

    yol = faz2_excel_yaz(forecast, master)
    print(f"\nExcel: {yol}")

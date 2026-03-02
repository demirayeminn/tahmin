"""
_faz3_excel.py — Faz 3 Excel çıktısı (T4 + T6 + T7).
Bu modül sadece faz3 çıktısı için kullanılır; src/export.py'nin genişlemesi.
"""
import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.config import DATA_OUTPUTS_DIR, LOG_FORMAT, LOG_LEVEL
from src.reorder import faz3_reorder_excel_sayfa
from src.anomaly import t6_turkce, t7_turkce

logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT)
log = logging.getLogger(__name__)

_HEADER_FILL = "1565C0"

_RISK_FILL = {
    "Kırmızı": "FFCDD2",
    "Turuncu": "FFE0B2",
    "Yeşil":   "C8E6C9",
    "Gri":     "EEEEEE",
}

_OLAY_FILL = {
    "spike":            "FFCDD2",
    "dusus":            "F3E5F5",
    "kampanya_benzeri": "FFF9C4",
    "olasi_stok_out":   "FFCCBC",
    "trend_artis":      "C8E6C9",
    "trend_dusus":      "B3E5FC",
}


def _format_ws(ws, renk_kolon: str | None = None, renk_map: dict | None = None):
    """Başlık stili + otomatik sütun genişliği + freeze."""
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if renk_kolon and renk_map:
        # Başlık satırından sütun index'i bul
        baslik_satirı = [cell.value for cell in ws[1]]
        if renk_kolon in baslik_satirı:
            kolon_idx = baslik_satirı.index(renk_kolon)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                deger = row[kolon_idx].value
                renk = renk_map.get(str(deger), "FFFFFF")
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=renk)

    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"


def faz3_excel_yaz(
    reorder: pd.DataFrame,
    events: pd.DataFrame,
    calendar: pd.DataFrame,
) -> Path:
    """
    Faz 3 sonuçlarını tek Excel dosyasına yazar.

    Sayfalar:
      T4_SatinAlma    — satın alma önerisi
      T6_Anomali      — anomali/olay tablosu
      T7_AykirıTakvim — aykırı günler takvimi

    Args:
        reorder:  satın alma önerisi DataFrame.
        events:   anomali olayları DataFrame.
        calendar: aykırı günler takvimi DataFrame.

    Returns:
        Yazılan dosyanın Path nesnesi.
    """
    DATA_OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    cikti = DATA_OUTPUTS_DIR / "faz3_satin_alma_anomali.xlsx"

    t4 = faz3_reorder_excel_sayfa(reorder)
    t6 = t6_turkce(events)
    t7 = t7_turkce(calendar)

    log.info("Faz 3 Excel yazılıyor: %s", cikti)
    with pd.ExcelWriter(cikti, engine="openpyxl") as writer:
        t4.to_excel(writer, sheet_name="T4_SatinAlma", index=False)
        t6.to_excel(writer, sheet_name="T6_Anomali", index=False)
        t7.to_excel(writer, sheet_name="T7_AykirıTakvim", index=False)

        wb = writer.book
        _format_ws(wb["T4_SatinAlma"], renk_kolon="Risk Bandı", renk_map=_RISK_FILL)
        _format_ws(wb["T6_Anomali"], renk_kolon="Olay Tipi", renk_map=_OLAY_FILL)
        _format_ws(wb["T7_AykirıTakvim"])

    log.info(
        "Faz 3 Excel kaydedildi: %s (T4:%d, T6:%d, T7:%d satır)",
        cikti, len(t4), len(t6), len(t7),
    )
    return cikti

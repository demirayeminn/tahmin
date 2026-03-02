"""
basket_analysis.py — Birlikte satış (Market Basket Analysis).

REF-05:
  Sepet    = Sipariş Numarası (aynı siparis_no → bir sepet)
  Algoritma: FP-Growth (mlxtend) — Apriori'den daha hızlı
  Min support   : BASKET_MIN_SUPPORT   (config'den, varsayılan 0.005)
  Min confidence: BASKET_MIN_CONFIDENCE (config'den, varsayılan 0.10)
  Min lift      : BASKET_MIN_LIFT       (config'den, varsayılan 1.0)

Aksiyon sütunu:
  lift > 3 ve confidence > 0.5 → "bundle_onerisi"
  lift > 2                     → "beraber_stokla"
  diğer                        → "cross_sell"

Çıktı (T5): urun_a_barkod, urun_b_barkod, urun_a_ad, urun_b_ad,
            support, confidence, lift, sepet_sayisi, aksiyon
"""

import logging
import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.config import (
    BASKET_MIN_CONFIDENCE,
    BASKET_MIN_LIFT,
    BASKET_MIN_SUPPORT,
    DATA_OUTPUTS_DIR,
    DATA_PROCESSED_DIR,
    LOG_FORMAT,
    LOG_LEVEL,
)

logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT)
log = logging.getLogger(__name__)

BASKET_RULES_PARQUET = DATA_PROCESSED_DIR / "basket_rules.parquet"


# ─── İşlem Matrisini Oluşturma ────────────────────────────────────────────────

def _islem_matrisi_olustur(clean_orders: pd.DataFrame) -> pd.DataFrame:
    """
    Sipariş verilerinden ikili (0/1) işlem matrisi üretir.
    Satır = siparis_no, Sütun = barkod.

    Sadece birden fazla ürün içeren sepetler dahil edilir
    (tek ürünlü sepetler birlikte satış analizi için anlamsız).
    """
    # Sadece normal (anormal_adet=False) kayıtlar
    normal = clean_orders[~clean_orders["anormal_adet"]].copy()

    # Sipariş başına ürün sayısı
    sepet_boyutu = normal.groupby("siparis_no")["barkod"].nunique()
    coklu_sepetler = sepet_boyutu[sepet_boyutu > 1].index
    normal = normal[normal["siparis_no"].isin(coklu_sepetler)]

    log.info(
        "Sepet analizi: toplam sipariş=%d | çoklu ürünlü=%d | tekil (atlandı)=%d",
        len(sepet_boyutu),
        len(coklu_sepetler),
        len(sepet_boyutu) - len(coklu_sepetler),
    )

    if normal.empty:
        raise ValueError("Çoklu ürünlü sepet bulunamadı.")

    # Pivot: siparis_no × barkod → 1 (satın alındı) / 0
    pivot = (
        normal.groupby(["siparis_no", "barkod"])["adet"]
        .sum()
        .gt(0)
        .unstack(fill_value=False)
        .astype(bool)
    )

    log.info(
        "İşlem matrisi: %d sepet × %d ürün",
        pivot.shape[0],
        pivot.shape[1],
    )
    return pivot


# ─── Kural Üretimi ────────────────────────────────────────────────────────────

def _kurallar_uret(pivot: pd.DataFrame, toplam_sepet: int) -> pd.DataFrame:
    """
    FP-Growth ile sık öğe kümeleri, ardından birliktelik kuralları üretir.

    Args:
        pivot: İkili işlem matrisi.
        toplam_sepet: Tüm sepet sayısı (support hesabı için).

    Returns:
        Filtrelenmiş birliktelik kuralları DataFrame.
    """
    from mlxtend.frequent_patterns import association_rules, fpgrowth

    log.info(
        "FP-Growth çalışıyor. Min support=%.4f, matris=%s",
        BASKET_MIN_SUPPORT,
        pivot.shape,
    )

    sikli_kumeler = fpgrowth(
        pivot,
        min_support=BASKET_MIN_SUPPORT,
        use_colnames=True,
        max_len=2,          # Sadece çift kurallar (A→B); üçlü+ atlanır
    )

    if sikli_kumeler.empty:
        log.warning("Sık öğe kümesi bulunamadı. Min support düşürülmeli?")
        return pd.DataFrame()

    log.info("Sık öğe kümesi sayısı: %d", len(sikli_kumeler))

    kurallar = association_rules(
        sikli_kumeler,
        metric="confidence",
        min_threshold=BASKET_MIN_CONFIDENCE,
        num_itemsets=len(sikli_kumeler),
    )

    # Lift filtresi
    kurallar = kurallar[kurallar["lift"] >= BASKET_MIN_LIFT]

    # Sadece ikili kurallar (antecedent ve consequent tek öğe)
    kurallar = kurallar[
        (kurallar["antecedents"].apply(len) == 1)
        & (kurallar["consequents"].apply(len) == 1)
    ]

    # Sepet sayısı: support × toplam_sepet
    kurallar["sepet_sayisi"] = (kurallar["support"] * toplam_sepet).round().astype(int)

    log.info(
        "Kural sayısı: %d (support≥%.3f, confidence≥%.2f, lift≥%.1f)",
        len(kurallar),
        BASKET_MIN_SUPPORT,
        BASKET_MIN_CONFIDENCE,
        BASKET_MIN_LIFT,
    )
    return kurallar


def _aksiyon_belirle(lift: float, confidence: float) -> str:
    """Lift ve confidence değerlerine göre iş aksiyonu döndür."""
    if lift > 3 and confidence > 0.5:
        return "bundle_onerisi"
    elif lift > 2:
        return "beraber_stokla"
    return "cross_sell"


def _t5_formatına_cevir(
    kurallar: pd.DataFrame,
    ad_map: dict[str, str],
) -> pd.DataFrame:
    """
    mlxtend kural DataFrame'ini T5 formatına çevirir.

    Args:
        kurallar: association_rules() çıktısı.
        ad_map:   barkod → kanonik_ad sözlüğü.

    Returns:
        T5 formatında DataFrame.
    """
    satirlar = []
    for _, satir in kurallar.iterrows():
        a_barkod = next(iter(satir["antecedents"]))
        b_barkod = next(iter(satir["consequents"]))
        satirlar.append({
            "urun_a_barkod":  a_barkod,
            "urun_b_barkod":  b_barkod,
            "urun_a_ad":      ad_map.get(a_barkod, a_barkod),
            "urun_b_ad":      ad_map.get(b_barkod, b_barkod),
            "support":        round(float(satir["support"]), 6),
            "confidence":     round(float(satir["confidence"]), 4),
            "lift":           round(float(satir["lift"]), 4),
            "sepet_sayisi":   int(satir["sepet_sayisi"]),
            "aksiyon":        _aksiyon_belirle(float(satir["lift"]), float(satir["confidence"])),
        })

    df = pd.DataFrame(satirlar)
    df = df.sort_values("lift", ascending=False).reset_index(drop=True)
    return df


# ─── Ana Fonksiyon ────────────────────────────────────────────────────────────

def birlikte_satis_analizi(
    clean_orders: pd.DataFrame,
    master: pd.DataFrame,
    kaydet: bool = True,
) -> pd.DataFrame:
    """
    Birlikte satış analizi çalıştırır ve T5 DataFrame döndürür.

    Args:
        clean_orders: Temiz sipariş DataFrame.
        master:       Ürün master (kanonik_ad için).
        kaydet:       True ise parquet kaydeder.

    Returns:
        T5 formatında birlikte satış kuralları DataFrame.
    """
    ad_map = master.set_index("barkod")["kanonik_ad"].to_dict()

    pivot = _islem_matrisi_olustur(clean_orders)
    toplam_sepet = len(pivot)

    kurallar = _kurallar_uret(pivot, toplam_sepet)

    if kurallar.empty:
        log.warning("Kural üretilemedi.")
        return pd.DataFrame(columns=[
            "urun_a_barkod", "urun_b_barkod", "urun_a_ad", "urun_b_ad",
            "support", "confidence", "lift", "sepet_sayisi", "aksiyon",
        ])

    t5 = _t5_formatına_cevir(kurallar, ad_map)

    aksiyon_dagilimi = t5["aksiyon"].value_counts().to_dict()
    log.info(
        "T5 hazır. Kural: %d | Aksiyon dağılımı: %s",
        len(t5), aksiyon_dagilimi,
    )

    if kaydet:
        DATA_PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
        t5.to_parquet(BASKET_RULES_PARQUET, index=False)
        log.info("Kaydedildi: %s", BASKET_RULES_PARQUET)

    return t5


# ─── Top-N Öneri Yardımcısı ───────────────────────────────────────────────────

def urun_icin_onerileri_getir(
    barkod: str,
    t5: pd.DataFrame,
    n: int = 10,
) -> pd.DataFrame:
    """
    Belirtilen ürünle birlikte en sık satılan N ürünü lift sıralı döndürür.
    Dashboard 03_birlikte_satis.py tarafından kullanılacak.
    """
    mask = (t5["urun_a_barkod"] == barkod) | (t5["urun_b_barkod"] == barkod)
    sonuc = t5[mask].copy()
    # Hedef ürün her zaman "A" tarafında olsun
    ters_mask = sonuc["urun_b_barkod"] == barkod
    sonuc.loc[ters_mask, ["urun_a_barkod", "urun_b_barkod"]] = (
        sonuc.loc[ters_mask, ["urun_b_barkod", "urun_a_barkod"]].values
    )
    sonuc.loc[ters_mask, ["urun_a_ad", "urun_b_ad"]] = (
        sonuc.loc[ters_mask, ["urun_b_ad", "urun_a_ad"]].values
    )
    return sonuc.sort_values("lift", ascending=False).head(n)


# ─── Excel Yazıcı ─────────────────────────────────────────────────────────────

def faz4_excel_yaz(t5: pd.DataFrame) -> Path:
    """
    Faz 4 T5 tablosunu formatlı Excel'e yazar.

    Sayfalar:
      T5_BirlikeSatis — tüm kurallar (lift azalan)
      Aksiyon_Ozet   — aksiyon bazında özet
      TopLift_20     — en yüksek lift'li 20 kural

    Args:
        t5: birlikte_satis_analizi() çıktısı.

    Returns:
        Yazılan dosyanın Path nesnesi.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    DATA_OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    cikti = DATA_OUTPUTS_DIR / "faz4_birlikte_satis.xlsx"

    _AKSIYON_FILL = {
        "bundle_onerisi": "FFCDD2",   # kırmızı ton
        "beraber_stokla": "FFF9C4",   # sarı ton
        "cross_sell":     "C8E6C9",   # yeşil ton
    }
    _HEADER_FILL = "1565C0"

    def _format_ws(ws, aksiyon_kolon: str | None = None):
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if aksiyon_kolon:
            basliklar = [c.value for c in ws[1]]
            if aksiyon_kolon in basliklar:
                k_idx = basliklar.index(aksiyon_kolon)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    renk = _AKSIYON_FILL.get(str(row[k_idx].value), "FFFFFF")
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor=renk)
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col_cells
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 50)
        ws.freeze_panes = "A2"

    # Türkçe başlıklar
    t5_tr = t5.rename(columns={
        "urun_a_barkod": "Ürün A Barkod",
        "urun_b_barkod": "Ürün B Barkod",
        "urun_a_ad": "Ürün A Adı",
        "urun_b_ad": "Ürün B Adı",
        "support": "Destek (Support)",
        "confidence": "Güven (Confidence)",
        "lift": "Lift",
        "sepet_sayisi": "Sepet Sayısı",
        "aksiyon": "Aksiyon",
    })

    # Aksiyon özeti
    aksiyon_ozet = (
        t5.groupby("aksiyon")
        .agg(
            kural_sayisi=("lift", "count"),
            ort_lift=("lift", "mean"),
            ort_confidence=("confidence", "mean"),
            ort_sepet=("sepet_sayisi", "mean"),
        )
        .round(3)
        .reset_index()
        .rename(columns={
            "aksiyon": "Aksiyon",
            "kural_sayisi": "Kural Sayısı",
            "ort_lift": "Ort. Lift",
            "ort_confidence": "Ort. Güven",
            "ort_sepet": "Ort. Sepet Sayısı",
        })
    )

    top20 = t5_tr.head(20)

    log.info("Faz 4 Excel yazılıyor: %s", cikti)
    with pd.ExcelWriter(cikti, engine="openpyxl") as writer:
        t5_tr.to_excel(writer, sheet_name="T5_BirlikeSatis", index=False)
        aksiyon_ozet.to_excel(writer, sheet_name="Aksiyon_Ozet", index=False)
        top20.to_excel(writer, sheet_name="TopLift_20", index=False)

        wb = writer.book
        _format_ws(wb["T5_BirlikeSatis"], aksiyon_kolon="Aksiyon")
        _format_ws(wb["Aksiyon_Ozet"])
        _format_ws(wb["TopLift_20"], aksiyon_kolon="Aksiyon")

    log.info("Faz 4 Excel kaydedildi: %s (%d kural)", cikti, len(t5))
    return cikti


# ─── Script olarak çalıştırma ─────────────────────────────────────────────────
if __name__ == "__main__":
    clean_orders = pd.read_parquet(DATA_PROCESSED_DIR / "clean_orders.parquet")
    master = pd.read_parquet(DATA_PROCESSED_DIR / "product_master.parquet")

    t5 = birlikte_satis_analizi(clean_orders, master)
    print(f"\nT5 kural sayısı : {len(t5)}")
    if not t5.empty:
        print(f"Aksiyon dağılımı:\n{t5['aksiyon'].value_counts().to_string()}")
        print(f"\nEn yüksek lift 5 kural:")
        print(t5[["urun_a_ad", "urun_b_ad", "support", "confidence", "lift", "aksiyon"]].head(5).to_string())
        yol = faz4_excel_yaz(t5)
        print(f"\nExcel: {yol}")

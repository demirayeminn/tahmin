"""
export.py — Formatlı Excel çıktısı üretir.

Faz 1 çıktısı: data/outputs/faz1_master_ve_talep.xlsx
  Sayfa 1 → T1: Ürün Master
  Sayfa 2 → T2: Günlük Talep (özet: ilk 50K satır)
"""

import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.config import (
    DATA_OUTPUTS_DIR,
    FAZ1_EXCEL,
    LOG_FORMAT,
    LOG_LEVEL,
    RENK_PALETI,
)

logging.basicConfig(level=LOG_LEVEL, format=LOG_FORMAT)
log = logging.getLogger(__name__)


# ─── Yardımcı ─────────────────────────────────────────────────────────────────

_ABC_FILL: dict[str, str] = {
    "A": "FFCDD2",   # açık kırmızı
    "B": "FFE0B2",   # açık turuncu
    "C": "C8E6C9",   # açık yeşil
}

_HEADER_FILL = "1565C0"   # koyu mavi
_HEADER_FONT_COLOR = "FFFFFF"


def _format_sayfa(ws, df: pd.DataFrame, sayfa_adi: str) -> None:
    """
    Verilen openpyxl worksheet'e başlık formatı + otomatik sütun genişliği uygula.
    """
    # Başlık satırı stili
    for cell in ws[1]:
        cell.font = Font(bold=True, color=_HEADER_FONT_COLOR, size=10)
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Satır renklendir (ABC sınıfına göre, varsa)
    if "abc_sinifi" in df.columns:
        abc_kolon_idx = df.columns.get_loc("abc_sinifi") + 1  # openpyxl 1-indexed
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            abc_val = row[abc_kolon_idx - 1].value
            renk = _ABC_FILL.get(str(abc_val), "FFFFFF")
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=renk)

    # Otomatik sütun genişliği
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    # Başlık satırını dondur
    ws.freeze_panes = "A2"
    log.debug("Sayfa formatlandı: %s", sayfa_adi)


# ─── Ana Fonksiyon ───────────────────────────────────────────────────────────

def faz1_excel_yaz(
    master: pd.DataFrame,
    daily: pd.DataFrame,
    cikti: Path = FAZ1_EXCEL,
    daily_max_satir: int = 50_000,
) -> Path:
    """
    Faz 1 sonuçlarını formatlı Excel dosyasına yazar.

    Sayfalar:
      T1_UrunMaster  — tüm ürün master tablosu
      T2_GunlukTalep — günlük talep (daily_max_satir ile sınırlı)

    Args:
        master: Ürün master DataFrame (T1).
        daily:  Günlük talep DataFrame (T2).
        cikti:  Çıktı dosya yolu.
        daily_max_satir: T2 sayfasına yazılacak maksimum satır.

    Returns:
        Yazılan dosyanın Path nesnesi.
    """
    DATA_OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

    # Türkçe sütun başlıkları
    master_tr = master.rename(
        columns={
            "barkod": "Barkod",
            "kanonik_ad": "Kanonik Ad",
            "alias_listesi": "Alias Listesi",
            "alias_sayisi": "Alias Sayısı",
            "ilk_satis_tarihi": "İlk Satış Tarihi",
            "son_satis_tarihi": "Son Satış Tarihi",
            "toplam_adet": "Toplam Adet",
            "toplam_siparis": "Toplam Sipariş",
            "gunluk_ort_adet": "Günlük Ort. Adet",
            "abc_sinifi": "ABC Sınıfı",
            "talep_tipi": "Talep Tipi",
            "aktif": "Aktif",
            "sku_saglik_skoru": "SKU Sağlık Skoru",
            "cv": "CV",
            "adi": "ADI",
        }
    )

    daily_tr = daily.rename(
        columns={
            "barkod": "Barkod",
            "tarih": "Tarih",
            "adet": "Adet",
            "kumulatif_adet": "Kümülatif Adet",
            "hafta_ici_mi": "Hafta İçi Mi",
            "hafta_no": "Hafta No",
            "ay": "Ay",
        }
    )

    daily_ozet = daily_tr.head(daily_max_satir)
    if len(daily_tr) > daily_max_satir:
        log.warning(
            "T2 sayfası %d satırla sınırlandırıldı (toplam: %d)",
            daily_max_satir,
            len(daily_tr),
        )

    log.info("Excel yazılıyor: %s", cikti)
    with pd.ExcelWriter(cikti, engine="openpyxl") as writer:
        master_tr.to_excel(writer, sheet_name="T1_UrunMaster", index=False)
        daily_ozet.to_excel(writer, sheet_name="T2_GunlukTalep", index=False)

        wb = writer.book
        _format_sayfa(wb["T1_UrunMaster"], master_tr, "T1_UrunMaster")
        _format_sayfa(wb["T2_GunlukTalep"], daily_ozet, "T2_GunlukTalep")

    log.info("Excel kaydedildi: %s (T1: %d satır, T2: %d satır)", cikti, len(master_tr), len(daily_ozet))
    return cikti


# ─── Script olarak çalıştırma ─────────────────────────────────────────────────
if __name__ == "__main__":
    from src.config import CLEAN_ORDERS_PARQUET, DAILY_DEMAND_PARQUET, PRODUCT_MASTER_PARQUET
    import pandas as pd

    master = pd.read_parquet(PRODUCT_MASTER_PARQUET)
    daily = pd.read_parquet(DAILY_DEMAND_PARQUET)
    yol = faz1_excel_yaz(master, daily)
    print(f"Excel çıktısı: {yol}")
